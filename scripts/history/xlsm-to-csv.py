#!/usr/bin/env python3
"""
Convert the league's history workbook into committed CSVs. Run once.

    python3 scripts/history/xlsm-to-csv.py [path/to/2025_FantasyWorld_History.xlsm]

## Why this exists as a committed script rather than a dependency

The app has six production dependencies and none of them parse Excel. Adding an
.xlsm parser would carry a permanent dependency for a job that runs once against
a file that is now terminal — from 2020 onward the league's record comes from the
Sleeper API, and this workbook will never be regenerated. So the conversion runs
here, offline, and the *output* is what the repo keeps.

This script never runs in CI or on Vercel. It is committed as provenance: it
records exactly how `data/history/*.csv` was produced, so a future reader can
re-derive them instead of trusting them.

## What it deliberately does NOT convert

The workbook's five derived sheets — `League Summary`, `Trends Over Time`,
`Member & Season Stats`, `Everyone vs Everyone`, `All-time Records` — are Excel
pivot and formula output, and every number on them is recomputed from base tables
by `src/lib/history.ts` instead. They are left out so that nobody can import them
by accident, because they carry errors the recomputation fixes: `League Summary`
reports Eric + Mark with 0 high-scorer weeks and $0 of the weekly side bet (a
lookup keying on `Eric` against data spelling it `Eric + Mark`; the true answer is
4 high, 2 low, +$20), and the hidden `All-time Records` sheet disagrees with the
dashboard about the all-time high score because the two were computed over
different eras.

The weekly sheets (`matchup_data`, `lineup_efficiency_weekly`,
`player_details_by_team`, `playoff_matchup_data`, `highlow_pts_per_week`) are also
skipped: Sleeper is the source of record for 2020+. `regular_season` is kept
anyway even though it overlaps 2020–2024, because the importer reconciles the
Sleeper pull against it — a second, independent derivation of those five seasons
is the strongest check available, and it costs 140 rows.

`player_total_points` IS kept despite being a weekly-era sheet, for one narrow
reason: it is the only place a 2021–2024 auction pick can be joined to a Sleeper
player id. That join hits 640/640.
"""

import csv
import hashlib
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python3 -m pip install openpyxl")

DEFAULT_WORKBOOK = Path.home() / "Desktop/Fantasy/2025_FantasyWorld_History.xlsm"
OUT_DIR = Path(__file__).resolve().parents[2] / "data" / "history"

# sheet name -> output stem. Order is the order of the manifest.
SHEETS = {
    "member": "member",
    "regular_season": "regular_season",
    "playoffs_legacy": "playoffs_legacy",
    "draft_locations": "draft_locations",
    "auction_drafts": "auction_drafts",
    "player_total_points": "player_total_points",
}

# `win_history` holds two unrelated tables side by side: the 2011+ podium in
# columns A-E, and a 2006-2010 champions-only block parked in columns L-O. They
# are split here rather than downstream, because a single CSV with two grains is
# a trap for whoever reads it next.
WIN_HISTORY_MAIN = ["championship_year", "place", "member_id", "member", "money_won"]
WIN_HISTORY_LEGACY = ["championship_year", "place", "member", "money_won"]


def cell(value):
    """Excel value -> CSV text, without inventing precision or losing it."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        # str() on a float is the shortest representation that round-trips, so
        # the sheet's own 2.440000000000012 survives intact rather than being
        # tidied into 2.44 -- if the source is imprecise, that is a fact about
        # the source and the importer's tolerances are set knowing it.
        return repr(value) if value != int(value) else str(int(value))
    return str(value).strip()


def write_csv(path, header, rows):
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)
    return hashlib.sha256(path.read_bytes()).hexdigest()


def dump_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = [cell(c) for c in rows[0]]
    # Trailing empty columns are an Excel artifact of the used-range, not data.
    while header and header[-1] == "":
        header.pop()
    width = len(header)
    body = []
    for raw in rows[1:]:
        vals = [cell(c) for c in raw[:width]]
        if not any(vals):
            continue
        body.append(vals)
    return header, body


def dump_win_history(ws):
    """Split the sheet's two side-by-side tables. Columns A-E and L-O."""
    rows = list(ws.iter_rows(values_only=True))[1:]
    main, legacy = [], []
    for raw in rows:
        left = [cell(c) for c in raw[0:5]]
        if any(left):
            main.append(left)
        # The legacy block starts at column L (index 11) and has no member_id --
        # 2006-2010 predates the membership record, so those champions are a name
        # and nothing more. money_won is a literal "-" meaning *unknown*, which
        # is not the same as the real 0 recorded for 2011-2013.
        right = [cell(c) for c in raw[11:15]]
        if any(right):
            legacy.append(right)
    return main, legacy


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not src.exists():
        sys.exit(f"workbook not found: {src}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    manifest = {"source": src.name, "files": {}}

    for sheet, stem in SHEETS.items():
        if sheet not in wb.sheetnames:
            sys.exit(f"sheet {sheet!r} is missing from {src.name}")
        header, body = dump_sheet(wb[sheet])
        path = OUT_DIR / f"{stem}.csv"
        digest = write_csv(path, header, body)
        manifest["files"][f"{stem}.csv"] = {
            "sheet": sheet,
            "rows": len(body),
            "columns": header,
            "sha256": digest,
        }
        print(f"  {stem}.csv  {len(body):>6} rows  {len(header)} cols")

    main_rows, legacy_rows = dump_win_history(wb["win_history"])
    for stem, header, rows in (
        ("win_history", WIN_HISTORY_MAIN, main_rows),
        ("legacy_champions", WIN_HISTORY_LEGACY, legacy_rows),
    ):
        path = OUT_DIR / f"{stem}.csv"
        digest = write_csv(path, header, rows)
        manifest["files"][f"{stem}.csv"] = {
            "sheet": "win_history",
            "rows": len(rows),
            "columns": header,
            "sha256": digest,
        }
        print(f"  {stem}.csv  {len(rows):>6} rows  {len(header)} cols")

    (OUT_DIR / "MANIFEST.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"\nwrote {len(manifest['files'])} files + MANIFEST.json to {OUT_DIR}")


if __name__ == "__main__":
    main()
